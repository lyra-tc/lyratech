import csv
import io
import re
import unicodedata

from openpyxl import Workbook, load_workbook

LEAD_FIELDS = (
    "name", "email", "phone", "company",
    "industry", "address", "service", "message",
)

TEMPLATE_HEADERS = [
    "Nombre", "Correo", "Teléfono", "Empresa",
    "Giro de la empresa", "Dirección", "Servicio", "Mensaje",
]

_HEADER_TO_FIELD = {
    "nombre": "name",
    "correo": "email",
    "email": "email",
    "telefono": "phone",
    "empresa": "company",
    "giro de la empresa": "industry",
    "giro": "industry",
    "direccion": "address",
    "servicio": "service",
    "mensaje": "message",
}

_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")

_MAX_DATA_ROWS = 2000

_MAX_LEN = {
    "name": 255, "email": 255, "phone": 50, "company": 255,
    "industry": 120, "address": 255, "service": 100,
}
_FIELD_LABEL = {
    "name": "Nombre", "email": "Correo", "phone": "Teléfono", "company": "Empresa",
    "industry": "Giro de la empresa", "address": "Dirección", "service": "Servicio",
}


class LeadImportError(Exception):
    """Raised for an unreadable / unsupported / empty upload."""


def _normalize_header(value) -> str:
    s = unicodedata.normalize("NFKD", str(value or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return " ".join(s.lower().split())


def _rows_from_grid(grid: list[list], filename: str) -> list[dict]:
    if not grid:
        raise LeadImportError(f"«{filename}» está vacío.")
    header = grid[0]
    field_by_index: dict[int, str] = {}
    used: set[str] = set()
    for idx, cell in enumerate(header):
        field = _HEADER_TO_FIELD.get(_normalize_header(cell))
        if field and field not in used:
            field_by_index[idx] = field
            used.add(field)
    if not field_by_index:
        raise LeadImportError(
            f"«{filename}»: no se reconoció ningún encabezado. Usa la plantilla."
        )

    rows: list[dict] = []
    for i, cells in enumerate(grid[1:], start=2):
        row = {f: "" for f in LEAD_FIELDS}
        has_value = False
        for idx, field in field_by_index.items():
            if idx < len(cells):
                raw = cells[idx]
                val = "" if raw is None else str(raw).strip()
                row[field] = val
                if val:
                    has_value = True
        if not has_value:
            continue
        row["_file"] = filename
        row["_row"] = i
        rows.append(row)
        if len(rows) > _MAX_DATA_ROWS:
            raise LeadImportError(
                f"«{filename}»: demasiadas filas (máximo {_MAX_DATA_ROWS})."
            )

    if not rows:
        raise LeadImportError(f"«{filename}» no tiene filas de datos.")
    return rows


def parse_upload(filename: str, content: bytes) -> list[dict]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        try:
            try:
                text = content.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = content.decode("latin-1")
            grid = [list(r) for r in csv.reader(io.StringIO(text))]
        except (csv.Error, UnicodeError) as exc:
            raise LeadImportError(
                f"«{filename}»: no se pudo leer como CSV."
            ) from exc
    elif name.endswith(".xlsx"):
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise LeadImportError(
                f"«{filename}»: no se pudo leer. ¿Es un .xlsx válido?"
            ) from exc
        try:
            ws = wb.active
            if ws is None:
                raise LeadImportError(f"«{filename}»: no tiene hoja de datos.")
            grid = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()
    elif name.endswith(".xls"):
        raise LeadImportError(
            f"«{filename}»: el formato .xls no es compatible. Guárdalo como .xlsx o .csv."
        )
    else:
        raise LeadImportError(f"«{filename}»: usa un archivo .csv o .xlsx.")
    return _rows_from_grid(grid, filename)


def validate_row(row: dict) -> str | None:
    if not row.get("name", "").strip():
        return "Falta el nombre"
    email = row.get("email", "").strip()
    if email and not _EMAIL_RE.match(email):
        return "Correo con formato inválido"
    for field, limit in _MAX_LEN.items():
        if len(row.get(field, "")) > limit:
            return f"{_FIELD_LABEL[field]} excede {limit} caracteres"
    return None


def plan_import(
    rows: list[dict],
    existing_emails: set[str],
    existing_phones: set[str],
) -> tuple[list[dict], list[tuple[dict, str]]]:
    to_insert: list[dict] = []
    skipped: list[tuple[dict, str]] = []
    seen_emails: set[str] = set()
    seen_phones: set[str] = set()
    for row in rows:
        reason = validate_row(row)
        if reason:
            skipped.append((row, reason))
            continue
        email = row.get("email", "").strip().lower()
        phone = row.get("phone", "").strip()
        if email and (email in existing_emails or email in seen_emails):
            skipped.append((row, "Correo ya registrado en leads"))
            continue
        if not email and phone and (phone in existing_phones or phone in seen_phones):
            skipped.append((row, "Teléfono ya registrado en leads"))
            continue
        to_insert.append(row)
        if email:
            seen_emails.add(email)
        if phone:
            seen_phones.add(phone)
    return to_insert, skipped


def build_xlsx(headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(list(r))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
