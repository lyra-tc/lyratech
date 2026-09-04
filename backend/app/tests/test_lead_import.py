import io

import pytest
from openpyxl import Workbook, load_workbook

from ..core.lead_import import (
    LeadImportError,
    TEMPLATE_HEADERS,
    build_xlsx,
    normalize_phone,
    parse_upload,
    plan_import,
    validate_row,
)


def _csv(rows):
    lines = [",".join(TEMPLATE_HEADERS)] + [",".join(r) for r in rows]
    return ("\n".join(lines)).encode("utf-8")


def test_parse_csv_maps_columns():
    rows = parse_upload("x.csv", _csv([["Ada", "ada@x.com", "55", "ACME", "TI", "Calle 1", "precio-fijo", "hola"]]))
    assert rows[0]["name"] == "Ada"
    assert rows[0]["email"] == "ada@x.com"
    assert rows[0]["industry"] == "TI"
    assert rows[0]["address"] == "Calle 1"
    assert rows[0]["_row"] == 2


def test_parse_header_normalization():
    body = "CORREO,nombre,giro de la empresa,teléfono\nada@x.com,Ada,Retail,55".encode("utf-8")
    rows = parse_upload("x.csv", body)
    assert rows[0]["name"] == "Ada"
    assert rows[0]["email"] == "ada@x.com"
    assert rows[0]["industry"] == "Retail"
    assert rows[0]["phone"] == "55"


def test_parse_skips_blank_rows():
    rows = parse_upload("x.csv", _csv([["Ada", "a@x.com", "", "", "", "", "", ""], ["", "", "", "", "", "", "", ""]]))
    assert len(rows) == 1


def test_parse_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.append(TEMPLATE_HEADERS)
    ws.append(["Bob", "bob@x.com", "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    rows = parse_upload("x.xlsx", buf.getvalue())
    assert rows[0]["name"] == "Bob"


def test_parse_xls_rejected():
    with pytest.raises(LeadImportError):
        parse_upload("x.xls", b"whatever")


def test_parse_unknown_ext_rejected():
    with pytest.raises(LeadImportError):
        parse_upload("x.txt", b"a,b")


def test_parse_no_headers_rejected():
    with pytest.raises(LeadImportError):
        parse_upload("x.csv", b"foo,bar\n1,2")


def test_validate_row():
    assert validate_row({"name": "", "email": ""}) == "Falta el nombre"
    assert validate_row({"name": "A", "email": "nope"}) == "Correo con formato inválido"
    assert validate_row({"name": "A", "email": "a@x.com"}) is None
    assert validate_row({"name": "A", "email": ""}) is None


def test_validate_row_rejects_overlong_value():
    reason = validate_row({"name": "A" * 300, "email": ""})
    assert reason is not None
    assert "Nombre" in reason


def test_parse_binary_as_csv_rejected():
    garbage = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x01\x00\x00\x00\x01\x00"
    with pytest.raises(LeadImportError):
        parse_upload("x.csv", garbage)


def test_parse_too_many_rows_rejected():
    rows = [["User %d" % i, "u%d@x.com" % i] for i in range(2001)]
    body = ("\n".join(
        [",".join(TEMPLATE_HEADERS)] + [",".join(r) for r in rows]
    )).encode("utf-8")
    with pytest.raises(LeadImportError):
        parse_upload("x.csv", body)


def test_plan_dedupes_existing_email():
    rows = [{"name": "A", "email": "a@x.com", "phone": ""}]
    to_insert, skipped = plan_import(rows, {"a@x.com"}, set())
    assert to_insert == []
    assert skipped[0][1] == "Correo ya registrado en leads"


def test_normalize_phone_strips_formatting():
    assert normalize_phone("( 442) 223 7492") == "4422237492"
    assert normalize_phone("(442)223-7492") == "4422237492"
    assert normalize_phone("+52 442 223 7492") == "4422237492"
    assert normalize_phone("52 442 223 7492") == "4422237492"
    assert normalize_phone("442.223.7492") == "4422237492"
    assert normalize_phone("") == ""
    assert normalize_phone(None) == ""


def test_plan_dedupes_phone_ignoring_formatting():
    rows = [{"name": "A", "email": "", "phone": "(442) 223-7492"}]
    to_insert, skipped = plan_import(rows, set(), {"4422237492"})
    assert to_insert == []
    assert "Teléfono" in skipped[0][1]


def test_plan_within_batch_phone_dedupe_ignores_formatting():
    rows = [
        {"name": "A", "email": "", "phone": "442 223 7492"},
        {"name": "B", "email": "", "phone": "(442)223-7492"},
    ]
    to_insert, skipped = plan_import(rows, set(), set())
    assert len(to_insert) == 1
    assert len(skipped) == 1


def test_plan_dedupes_phone_only_when_no_email():
    rows = [{"name": "A", "email": "", "phone": "55"}]
    to_insert, skipped = plan_import(rows, set(), {"55"})
    assert to_insert == []
    assert "Teléfono" in skipped[0][1]


def test_plan_phone_dup_ignored_when_row_has_email():
    rows = [{"name": "A", "email": "a@x.com", "phone": "55"}]
    to_insert, _ = plan_import(rows, set(), {"55"})
    assert len(to_insert) == 1


def test_plan_within_batch_dedupe():
    rows = [
        {"name": "A", "email": "a@x.com", "phone": ""},
        {"name": "A2", "email": "A@X.COM", "phone": ""},
    ]
    to_insert, skipped = plan_import(rows, set(), set())
    assert len(to_insert) == 1
    assert len(skipped) == 1


def test_build_xlsx_roundtrip():
    data = build_xlsx(["H1", "H2"], [["a", "b"]])
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active
    assert [c.value for c in ws[1]] == ["H1", "H2"]
    assert [c.value for c in ws[2]] == ["a", "b"]


from ..core.lead_import import TEMPLATE_HEADERS as _TH


def _csv_bytes(rows):
    lines = [",".join(_TH)] + [",".join(r) for r in rows]
    return ("\n".join(lines)).encode("utf-8")


def _row(name="", email="", phone="", company="", industry="", address="", service="", message=""):
    return [name, email, phone, company, industry, address, service, message]


def test_import_requires_admin(client, non_admin_client):
    files = {"files": ("x.csv", _csv_bytes([_row("A", "a@x.com")]), "text/csv")}
    assert client.post("/api/leads/import", files=files).status_code == 401
    assert non_admin_client.post("/api/leads/import", files=files).status_code == 403
    assert client.get("/api/leads/import/template").status_code == 401


def test_template_download(auth_client):
    resp = auth_client.get("/api/leads/import/template")
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "plantilla-leads.xlsx" in resp.headers["content-disposition"]
    assert len(resp.content) > 0


def test_import_csv_inserts_rows(auth_client):
    files = {"files": ("leads.csv", _csv_bytes([
        _row("Ada", "ada@x.com", "1", "ACME", "TI", "Calle 1", "precio-fijo", "hola"),
        _row("Bob", "bob@x.com"),
    ]), "text/csv")}
    resp = auth_client.post("/api/leads/import", files=files)
    assert resp.status_code == 200
    body = resp.json()
    assert body["inserted"] == 2
    assert body["skipped_count"] == 0
    assert body["report_xlsx_base64"] is None
    listed = auth_client.get("/api/leads/").json()["items"]
    ada = next(x for x in listed if x["email"] == "ada@x.com")
    assert ada["industry"] == "TI"
    assert ada["address"] == "Calle 1"


def test_import_xlsx_inserts_rows(auth_client):
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(_TH)
    ws.append(_row("Xls User", "xls@x.com"))
    buf = io.BytesIO()
    wb.save(buf)
    files = {"files": ("leads.xlsx", buf.getvalue(),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = auth_client.post("/api/leads/import", files=files)
    assert resp.status_code == 200
    assert resp.json()["inserted"] == 1


def test_import_skips_duplicate_email(auth_client):
    auth_client.post("/api/leads/manual", json={"name": "Existing", "email": "dup@x.com"})
    files = {"files": ("l.csv", _csv_bytes([
        _row("New", "new@x.com"),
        _row("Dup", "DUP@x.com"),
    ]), "text/csv")}
    body = auth_client.post("/api/leads/import", files=files).json()
    assert body["inserted"] == 1
    assert body["skipped_count"] == 1
    assert body["skipped"][0]["reason"] == "Correo ya registrado en leads"
    assert body["skipped"][0]["row"] == 3
    assert body["report_xlsx_base64"] is not None


def test_import_skips_duplicate_phone_when_no_email(auth_client):
    auth_client.post("/api/leads/manual", json={"name": "Existing", "phone": "5551234"})
    files = {"files": ("l.csv", _csv_bytes([_row("Dup", "", "5551234")]), "text/csv")}
    body = auth_client.post("/api/leads/import", files=files).json()
    assert body["inserted"] == 0
    assert "Teléfono" in body["skipped"][0]["reason"]


def test_import_skips_duplicate_phone_different_format(auth_client):
    auth_client.post("/api/leads/manual", json={"name": "Existing", "phone": "(442) 223-7492"})
    files = {"files": ("l.csv", _csv_bytes([_row("Dup", "", "442 223 7492")]), "text/csv")}
    body = auth_client.post("/api/leads/import", files=files).json()
    assert body["inserted"] == 0
    assert "Teléfono" in body["skipped"][0]["reason"]


def test_import_skips_invalid_rows(auth_client):
    files = {"files": ("l.csv", _csv_bytes([
        _row("", "a@x.com"),
        _row("BadMail", "not-an-email"),
        _row("Good", "good@x.com"),
    ]), "text/csv")}
    body = auth_client.post("/api/leads/import", files=files).json()
    assert body["inserted"] == 1
    reasons = {s["reason"] for s in body["skipped"]}
    assert "Falta el nombre" in reasons
    assert any("inválido" in r for r in reasons)


def test_import_skips_within_batch_duplicates(auth_client):
    files = {"files": ("l.csv", _csv_bytes([
        _row("First", "same@x.com"),
        _row("Second", "same@x.com"),
    ]), "text/csv")}
    body = auth_client.post("/api/leads/import", files=files).json()
    assert body["inserted"] == 1
    assert body["skipped_count"] == 1


def test_import_skips_overlong_value(auth_client):
    files = {"files": ("l.csv", _csv_bytes([
        _row("Good", "good2@x.com"),
        _row("BigCo", "big@x.com", "", "C" * 300),
    ]), "text/csv")}
    body = auth_client.post("/api/leads/import", files=files).json()
    assert body["inserted"] == 1
    assert body["skipped_count"] == 1
    assert "Empresa" in body["skipped"][0]["reason"]


def test_import_rejects_binary_as_csv(auth_client):
    png = b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x01\x00\x00\x00\x01\x00"
    files = {"files": ("x.csv", png, "text/csv")}
    resp = auth_client.post("/api/leads/import", files=files)
    assert resp.status_code == 400


def test_import_rejects_xls(auth_client):
    files = {"files": ("old.xls", b"garbage", "application/vnd.ms-excel")}
    resp = auth_client.post("/api/leads/import", files=files)
    assert resp.status_code == 400
    assert ".xls" in resp.json()["detail"]


def test_import_no_notification(auth_client, monkeypatch):
    auth_client.post("/api/notifications/recipients", json={"email": "team@x.com"})
    called = []
    monkeypatch.setattr(
        "app.routers.leads.send_lead_notification_email",
        lambda *a, **k: called.append(1),
    )
    files = {"files": ("l.csv", _csv_bytes([_row("A", "a@x.com")]), "text/csv")}
    auth_client.post("/api/leads/import", files=files)
    assert called == []
